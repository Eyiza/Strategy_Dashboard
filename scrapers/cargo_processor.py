import pdfplumber
import pandas as pd
import os
import zipfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import warnings

warnings.filterwarnings('ignore')

# --- CONFIGURATION ---
JETTY_STATE_MAPPING = {
    'LAGOS': 'Lagos', 'APAPA': 'Lagos', 'TINCAN': 'Lagos', 'TIN CAN': 'Lagos',
    'ATLAS COVE': 'Lagos', 'IJEGUN': 'Lagos', 'KIRIKIRI': 'Lagos',
    'FOLAWIYO': 'Lagos', 'PPMC': 'Lagos', 'NACJ': 'Lagos', 'SBM': 'Lagos',
    'NOJ': 'Lagos', 'BOP': 'Lagos', 'FISHERY': 'Lagos', 'MARINA': 'Lagos',
    'PORT HARCOURT': 'Rivers', 'RIVERS': 'Rivers', 'ONNE': 'Rivers',
    'OKRIKA': 'Rivers', 'BONNY': 'Rivers', 'DAWES': 'Rivers', 'PHRC': 'Rivers',
    'WARRI': 'Delta', 'DELTA': 'Delta', 'ESCRAVOS': 'Delta', 'FORCADOS': 'Delta',
    'KOKO': 'Delta', 'BURUTU': 'Delta', 'BENETH': 'Delta', 'BENNETT': 'Delta',
    'OGHARA': 'Delta', 'WRPC': 'Delta',
    'CALABAR': 'Cross River',
    'IBENO': 'Akwa Ibom', 'EKET': 'Akwa Ibom', 'IBAKA': 'Akwa Ibom'
}

FOREIGN_KEYWORDS = ['GHANA', 'ABIDJAN', 'LOME', 'TOGO', 'COTONOU', 'BENIN', 'IVORY COAST']

# --- HELPERS ---
def clean_jetty_name(text):
    if not text: return ""
    text_str = str(text)
    split_index = text_str.lower().find(" lat")
    if split_index != -1: text_str = text_str[:split_index]
    return text_str.strip()

def is_foreign_entry(jetty_info):
    if not jetty_info: return False
    info_upper = str(jetty_info).upper()
    for keyword in FOREIGN_KEYWORDS:
        if keyword in info_upper: return True
    return False

def get_state_from_jetty(jetty_info):
    if not jetty_info: return ''
    jetty_info_str = str(jetty_info).upper()
    for key, state in JETTY_STATE_MAPPING.items():
        if key in jetty_info_str: return state
    return ''

def parse_date(date_string):
    if not date_string or str(date_string).strip() in ['-', '', 'None']: return None
    date_part = str(date_string).strip().upper()
    if 'DATE:' in date_part: date_part = date_part.split('DATE:')[1].strip()
    date_part = date_part.replace('SEPT', 'SEP') # Fix SEPT bug
    
    formats = ['%d-%b-%y', '%d-%b-%Y', '%d-%m-%Y', '%d %b-%y', '%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y']
    for fmt in formats:
        try: return datetime.strptime(date_part, fmt)
        except ValueError: continue
    return date_string

def split_bundled_row(entry_data):
    cargo = str(entry_data.get('Cargo', ''))
    if '/' not in cargo: return [entry_data]
    qty = str(entry_data.get('Quantity [MT]', ''))
    cargo_parts = [c.strip() for c in cargo.split('/')]
    qty_parts = [q.strip() for q in qty.split('/')]
    split_entries = []
    for i, cargo_item in enumerate(cargo_parts):
        new_row = entry_data.copy()
        new_row['Cargo'] = cargo_item
        new_row['Quantity [MT]'] = qty_parts[i] if i < len(qty_parts) else (qty_parts[-1] if qty_parts else qty)
        split_entries.append(new_row)
    return split_entries

# --- PARSING LOGIC ---
def is_jetty_row(row):
    non_none = [c for c in row if c and str(c).strip()]
    return len(non_none) == 1

def is_field_row(row):
    non_none = [str(c).strip().upper() for c in row if c and str(c).strip()]
    expected = ['POSITION', "SHIP'S NAME", 'CARGO', 'QTY', 'ARRVD', 'ETB', 'SAILED']
    if len(non_none) >= 4:
        matches = sum(1 for cell in non_none for exp in expected if exp in cell)
        return matches >= 3
    return False

def is_entry_row(row):
    non_none = [c for c in row if c and str(c).strip()]
    return (len(non_none) > 0 and not is_jetty_row(row) and not is_field_row(row))

def parse_pdf_to_excel(filepath, output_filepath):
    all_data = []
    date = None
    current_jetty = None
    fixed_headers = ['Date', 'State', 'Jetty Information', 'Position', "Ship's Name", 'Cargo', 'Quantity [MT]', 'ETA', 'ETB', 'Sailed [ETD]', 'Charterers/Receivers', 'Remarks']

    with pdfplumber.open(filepath) as pdf:
        pages = pdf.pages
        for page_index, page in enumerate(pages):
            # 1. Date (Page 1)
            if page_index == 0:
                 tables = page.extract_tables()
                 if tables and len(tables[0]) > 2:
                     row = tables[0][2]
                     if len(row) > 9 and row[9]: date = parse_date(str(row[9]))

            # 2. Jetty Info (Top of Page)
            if page_index > 0:
                tables = page.extract_tables()
                if tables and tables[0] and is_field_row(tables[0][0]):
                    page_text = page.extract_text()
                    if page_text:
                        for line in page_text.split('\n'):
                            if line.strip() and 'page' not in line.lower():
                                current_jetty = clean_jetty_name(line.strip())
                                break

            # 3. Rows
            tables = page.extract_tables()
            if tables:
                table = tables[0]
                rows_to_process = table[:-1] if page_index == len(pages) - 1 else table
                for row in rows_to_process:
                    if not row: continue
                    if is_jetty_row(row):
                        non_none = [str(c).strip() for c in row if c and str(c).strip()]
                        current_jetty = clean_jetty_name(non_none[0] if non_none else "")
                    elif is_entry_row(row) and current_jetty:
                        if is_foreign_entry(current_jetty): continue
                        # Process Row Logic
                        processed = [str(c).strip() if c else None for c in row]
                        # Find last valid index
                        last_idx = -1
                        for i in range(len(processed)-1, -1, -1):
                            if processed[i] is not None: 
                                last_idx = i
                                break
                        if last_idx == -1: continue
                        
                        data = processed[:last_idx+1]
                        if len(data) > 9: data = data[-9:]
                        entry = [(item if item else '-') for item in data]
                        while len(entry) < 9: entry.append('-')
                        
                        if entry[0] == '-' or entry[0].upper() == 'VACANT': continue
                        
                        # Swap Cargo/Qty fix
                        if (entry[2] is None or entry[2] in ['-', '']) and (entry[3] not in [None, '-', '']):
                            entry[2] = entry[3]
                            entry[3] = '-'
                            
                        base = {
                            'Date': date, 'State': get_state_from_jetty(current_jetty),
                            'Jetty Information': current_jetty, 'Position': entry[0],
                            "Ship's Name": entry[1], 'Cargo': entry[2], 'Quantity [MT]': entry[3],
                            'ETA': parse_date(entry[4]) if entry[4] != '-' else '-',
                            'ETB': parse_date(entry[5]) if entry[5] != '-' else '-',
                            'Sailed [ETD]': parse_date(entry[6]) if entry[6] != '-' else '-',
                            'Charterers/Receivers': entry[7], 'Remarks': entry[8]
                        }
                        all_data.extend(split_bundled_row(base))

    if all_data:
        df = pd.DataFrame(all_data)
        df = df.reindex(columns=fixed_headers, fill_value='')
        df.to_excel(output_filepath, index=False, engine='openpyxl')
        format_excel_file(output_filepath)
        return df
    return None

def format_excel_file(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    font = Font(name='Raleway', size=11)
    align = Alignment(horizontal='center', vertical='center')
    header_font = Font(name='Raleway', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            if cell.value is not None:
                cell.font = font
                cell.alignment = align
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                if isinstance(cell.value, datetime):
                    cell.number_format = 'd-mmm-yy'

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_len + 5, 60)
    wb.save(filepath)

# --- MAIN EXPORT FUNCTION ---
def process_cargo_files(uploaded_files, download_folder):
    """
    uploaded_files: List of FileStorage objects from Flask
    """
    temp_dir = os.path.join(download_folder, "temp_pdfs")
    if not os.path.exists(temp_dir): os.makedirs(temp_dir)
    
    processed_paths = []
    master_dfs = []
    
    # 1. Process Each File
    for file in uploaded_files:
        if not file.filename.lower().endswith('.pdf'): continue
        
        # Save PDF temporarily
        pdf_path = os.path.join(temp_dir, file.filename)
        file.save(pdf_path)
        
        # Output Path
        xlsx_name = "CLEANED_" + file.filename.replace('.pdf', '.xlsx')
        xlsx_path = os.path.join(download_folder, xlsx_name)
        
        # Run Parser
        try:
            df = parse_pdf_to_excel(pdf_path, xlsx_path)
            if df is not None:
                processed_paths.append(xlsx_path)
                master_dfs.append(df)
        except Exception as e:
            print(f"Error parsing {file.filename}: {e}")
        
        # Cleanup PDF
        os.remove(pdf_path)
        
    if not master_dfs:
        return None

    # 2. Create Master File
    master_filename = "MASTER_MERGED_CARGO_DATA.xlsx"
    master_path = os.path.join(download_folder, master_filename)
    
    master_df = pd.concat(master_dfs, ignore_index=True)
    master_df.to_excel(master_path, index=False, engine='openpyxl')
    format_excel_file(master_path)
    
    # 3. Zip Everything (Master + Individual Cleaned Files)
    zip_filename = "Cargo_Analysis_Results.zip"
    zip_path = os.path.join(download_folder, zip_filename)
    
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        zipf.write(master_path, master_filename)
        for p in processed_paths:
            zipf.write(p, os.path.basename(p))
            os.remove(p) # Remove individual excel files after zipping to save space
    
    os.remove(master_path) # Remove master excel after zipping
    return zip_filename